#!/usr/bin/env python3
# -*- coding: utf-8 -*-

#1.0 目前支持的功能，从已翻译的工程中提取中文，和已翻译的泰文、英文和香港繁体字，并形成一份翻译汇总
#1.1 支持找寻所有本地化文件，不需要自己枚举

import math
import os
import re
import sys
import openpyxl
import time
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font,colors,Alignment

#用来存放其他语言的key-value字典，组成一个数组,例：
#{
# en:{
#    {wehcat:微聊},
#    {other:其他},
#   },
# th:{
#    {wehcat:微聊},
#    {other:其他},
#   },
#}
global_translate_dict = {}
file_lproj_zh_hans = 'zh-Hans'
file_iOS = 'iOS翻译汇总'
#是否更新到总表
merge_to_totalExcel = False
#是否需要生成一份未翻译的表格，供翻译人员翻译新的语种
generative_new_empty_sheet = False
#新表格名字
generative_new_empty_sheet_name = '翻译'
#这里可以自定义对应语言的sheetName，如果没有定义，默认就是【en翻译、th翻译、zh-Hant翻译】
global_sheetName_dict = {'en':'英文翻译','th':'泰文翻译','zh-Hant':'香港繁体翻译','id':'印尼翻译'}

localString_file_Baimingdan = ['BaiduLoc_iOSSDK_Libs_No_Bitcode_No_IDFA_v1.1','RETableViewManager','AMap3DMap-NO-IDFA','AMapFoundation-NO-IDFA','AMapLocation-NO-IDFA','AMapSearch-NO-IDFA','BaiduMapKit','YYImage','UMCCommon','UMCAnalytics','PLPlayerKit','EKOToast','EKOThird','Alipay','EKOSDK','EKORealTimeCallLib','EKOLogger','GoogleMaps','GooglePlaces'] #路径白名单，路径包含这个的，不会处理

# 修改excel示例
def modify_excel_file():
    wb = load_workbook('iOS翻译汇总.xlsx')
    #根据[英文翻译]这个sheet名字来获取该sheet
    ws = wb['英文翻译']
    print(ws['A1'].value)
    ws['A1'] = 'tttttttttt'
    wb.save('iOS翻译汇总1.xlsx')

def write_to_excel_file():
    handle_excel_file()

def handle_excel_file():
    
    chinese_lanaguage_dict = global_translate_dict[file_lproj_zh_hans]
    if len(chinese_lanaguage_dict) == 0:
        print('没有找到本地化中文文件？？？？？？？')
        return
    global_translate_dict.pop(file_lproj_zh_hans)
    #将数据写入excel
    #1.工作表相关,需要导入from openpyxl import Workbook
    #这样就新建了一个新的工作表（只是还没被保存）,默认表名为sheet
    wb = Workbook()
    #删除某一张表
    wb.remove(wb.active)
    #最重要的一步是要保存
    currentTime = time.strftime('%Y年%m月%d日%H时%M分%S秒',time.localtime())
    fileName = os.path.join(os.getcwd(),file_iOS + currentTime + '.xlsx')
###########################################################################################
    bold_itatic_24_font = Font(color = colors.RED)
    bold_itatic_chines_font = Font(color = colors.BLUE)
    for lanaguage_key in global_translate_dict.keys():
        index = list(global_translate_dict.keys()).index(lanaguage_key)
        translate_dict = global_translate_dict[lanaguage_key]
        sheetName = global_sheetName_dict.get(lanaguage_key,generative_new_empty_sheet_name)
        wb.create_sheet(sheetName,index = index)
        sheet = wb[sheetName]
        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 500
        count = 1
        
        flagIndex_chinese = 'A'+ str(count)
        sheetIndex = 'B'+ str(count)
        sheet.row_dimensions[count].height = 20
        
        sheet[sheetIndex].font = bold_itatic_24_font
        sheet[flagIndex_chinese].font = bold_itatic_24_font
        
        sheet[flagIndex_chinese] = '翻译完成YES/NO'
        sheet[sheetIndex] = '注意：如果中间带有‘/’或者%@或者%d这种字符串的，翻译过程中请不要去掉，这个是程序中需要替换的词语，比如【%@的手表】可以理解为【我的手表】'
        for key in chinese_lanaguage_dict.keys():
            count  = count + 1
            flagIndex_chinese = 'A'+ str(count)
            sheetIndex = 'B'+ str(count)
            sheet.row_dimensions[count].height = 20
            sheet[sheetIndex].font = bold_itatic_chines_font
            sheet[sheetIndex] = chinese_lanaguage_dict[key]
            count  = count + 1
            sheetIndex = 'B'+ str(count)
            sheet.row_dimensions[count].height = 20
            if key in translate_dict.keys():
                sheet[flagIndex_chinese] = 'YES'
                flagIndex_chinese = 'A'+ str(count)
                sheet[flagIndex_chinese] = 'YES'
                sheet[sheetIndex] = translate_dict[key]
            else:
                if lanaguage_key != 'other':
                    print('该中文没有翻译成【%s】：%s=%s' % (sheetName,key,chinese_lanaguage_dict[key]))
                sheet[flagIndex_chinese] = 'NO'
                flagIndex_chinese = 'A'+ str(count)
                sheet[flagIndex_chinese] = 'NO'
                sheet[sheetIndex] = ''
                
    wb.save(r'%s' % fileName)

def search_all_strings_file(filePath):
    #搜索各类本地化文件，找出key和value
    startSearch_all_strings_file(filePath)

def startSearch_all_strings_file(filePath,encodeingType = 'utf-8'):
    for dir in os.listdir(filePath):
        if xtc_is_localString_BaiMingDan(dir):
            continue
        path = os.path.join(filePath,dir)
        if os.path.isdir(path):
            #因为bundle里面的文件用【utf-8】读取会越界，所以需要更改读取方式为【utf-16LE】
            if (os.path.splitext(path)[1] in ['.bundle','.framework']):
                if path.split('/')[-1] in ['emojiBase.bundle']:
                    return
                print('手动确认下，这个bundle文件下的是不是需要整理：' + path)
                startSearch_all_strings_file(path,'utf-16LE')
            else:
                startSearch_all_strings_file(path,encodeingType)
        else:
            #1、获取指定的string文件路径
            if os.path.splitext(path)[1] in ['.strings']:
                print(path)
                abspath = os.path.abspath(path)
                #2、获取是什么语言
                language_lproj = abspath.split('/')[-2].split('.')[0]
                print(language_lproj)
                #2.1、base语言不作处理
                if language_lproj in ['Base']:
                    break
                language_dict = dict_language_with_key(language_lproj)
                #3、获取指定的string文件的key，value
                open_localString_project_file(path,language_lproj,language_dict,encodeingType)
                print('扫描该文件后，一共有%s个==========>>>%s模式' % (len(language_dict),language_lproj))

def open_localString_project_file(filePath,lproj,language_dict,encodeingType = 'utf-8'):
    repeatCount = 0
    lineNumber = 0
    count = 0
    with open(filePath,'r',encoding = encodeingType,) as f:
        for line in f.readlines():
            lineNumber = lineNumber + 1
            key = string_localString_key_line(line)
            string = string_localString_value_line(line)
            if lproj == file_lproj_zh_hans:
                if string is not None and len(string) > 0:
                    if string in language_dict.values():
                        repeatCount = repeatCount + 1
                        print('该字符串已经有了，不需要重新导入：第%s行:%s' % (lineNumber,string))
                    else:
                        language_dict[key] = string
                    count = count + 1
            else:
                if key is not None:
                    if key in language_dict.keys():
                        repeatCount = repeatCount + 1
                        print('该key已经有了，不需要重新导入：第%s行:%s' % (lineNumber,key))
                    else:
                        language_dict[key] = string
                    count = count + 1
        f.close()
    print('一共有%s个key,其中有%s个重复==========>>>%s模式' % (count,repeatCount,lproj))

def dict_language_with_key(key):
    dict_language = global_translate_dict.get(key,{})
    global_translate_dict[key] = dict_language
    return dict_language

# 获取中文部分（不包含双引号）：
# "StrBannerLagalHoliday" = "今日放假，不禁用";
#输出：今日放假，不禁用
def string_localString_value_line(line):
    array = re.findall(r'"(.*?)"',line)
    if len(array) == 1 or len(array) == 2:
        return array[-1]
    elif len(array) > 2:
        #1、先取出带双引号部分，非贪婪模式
        array = line.split('=')
        firstString = array[0]
        #2、去掉第一个带双引号部分
        line = line.replace(firstString,"")
        #3、以贪婪模式，获取双引号内部的内容
        array = re.findall(r'"(.+)"',line)
        return array[0]

# 获取key部分（不包含双引号）：
# "StrBannerLagalHoliday" = "今日放假，不禁用";
#输出：StrBannerLagalHoliday
def string_localString_key_line(line):
    array = re.findall(r'"(.+?)"',line)
    if len(array) >= 2:
        return array[0]
    elif len(array) == 1:
        # info.plst里面的文案
        array = re.findall(r'\w+',line)
        return array[0]
    else:
        return None

def xtc_is_localString_BaiMingDan(path):
    if path in localString_file_Baimingdan:
        return True
    return False

if __name__ == '__main__':
    
#    if input('是否需要更新到总表里面：Y/N:') == 'Y':
#        print('需要合并到总表里面')
#        merge_to_totalExcel = True
#    else:
#        print('不需要合并到总表里面')
    inputString = input('是否需要生成新的表格供翻译：Y/N:')
    if inputString == 'Y' or inputString == 'y':
        print('需要生成新的表格')
        generative_new_empty_sheet = True
        generative_new_empty_sheet_name = input('请输入新表格的名字:')
        global_translate_dict['other'] = {}
    else:
        print('不需要生成新的表格')

#1、搜索工程中的所有中文.string文件
    startpath = os.path.abspath(os.path.dirname(os.getcwd()))
    search_all_strings_file(startpath)
#2、写入excel文件
    write_to_excel_file()

