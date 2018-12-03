#!/usr/bin/env python3
# -*- coding: utf-8 -*-
#函数篇
import math
import os
import re
import sys
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font,colors,Alignment
#•python处理excel已经有大量包，主流代表有：
#•xlwings：简单强大，可替代VBA
#•openpyxl：简单易用，功能广泛
#•pandas：使用需要结合其他库，数据处理是pandas立身之本
#•win32com：不仅仅是excel，可以处理office;不过它相当于是 windows COM 的封装，新手使用起来略有些痛苦。
#•Xlsxwriter：丰富多样的特性，缺点是不能打开/修改已有文件，意味着使用 xlsxwriter 需要从零开始。
#•DataNitro：作为插件内嵌到excel中，可替代VBA，在excel中优雅的使用python
#•xlutils：结合xlrd/xlwt，老牌python包，需要注意的是你必须同时安装这三个库
import xlrd
import xlwt
from xlutils.copy import copy

def main(filePath):
    handle_excel_file()

def handle_excel_file():
    exl = openpyxl.load_workbook('翻译.xlsx')
    #获取所有sheet的名称
    sheetNameArray = exl.sheetnames
    print(sheetNameArray)
    if len(sheetNameArray) == 0:
        print("当前的excel无sheet，需要创建后，再运行python")
        return None
    #获取第一个sheet的名称
    firstSheetName = sheetNameArray[0]
    #根据sheet名获取sheet
    firstSheet = exl[firstSheetName]
    print(firstSheet.title)
    #获取当前正在显示的sheet，一般都是第一次个sheet
    activeSheet = exl.active
    print(activeSheet)

#获取某个单元格的值，观察excel发现也是先字母再数字的顺序，即先列再行
    b4 = activeSheet['B4']
    print(b4)
#分别返回(B,4) is None
    print(f'({b4.column},{b4.row}) is {b4.value}')
#除了用下表的方式获得，还可以用cell函数，换成数字，这个表示B4
    b4_too = activeSheet.cell(row = 4,column = 2)
    print(b4_too.value)
#cell还有一个属性coordinate,像b4这个单元格坐标是B4
    print(b4_too.coordinate)
#获取最大行和最大列
    print(activeSheet.max_row)
    print(activeSheet.max_column)
#获取行和列
# sheet.rows为生成器，里面是每一行的数据，每一行又由一个tuple包裹
# sheet.columns类似，不过里面是每个tuple是每一列的单元格
    #因为按行，所以返回A1，B1,C1这样的顺序
    print('打印每一行>>>>>>>>>>>>')
#    for row in activeSheet.rows:
#        for cell in row:
#            print(cell.value)
    print('打印每一列>>>>>>>>>>>>')
    #A1,A2,A3这样的顺序
#    for column in activeSheet.columns:
#        for cell in column:
#            print(cell.value)
#如果先要获得某行的数据，给其一个索引就可以了，因为sheet.rows是生成器类型，不能使用索引，转换成list之后再使用索引,list(sheet.rows)[2]这样就可以获得第三行的tuple对象
    print('第三行的数据>>>>>>>>>>>>')
#    for cell in list(activeSheet.rows)[2]:
#        print(cell.value)
# range(1,4)返回1，2，3
#返回某一个区域的数据，可以使用sheet['A1':'B3']返回一个tuple，该元组内部还是元组，返回A1为左上角，B3为右下角的区域
#    for row_cell in activeSheet['A1':'B3']:
#        for cell in row_cell:
#            print(cell)
###########################################################################################
########################建表这里有问题#######################################################
###########################################################################################
    #将数据写入excel
    #1.工作表相关,需要导入from openpyxl import Workbook
    #这样就新建了一个新的工作表（只是还没被保存）,默认表名为sheet
    wb = Workbook()
    #若要指定【只写模式】，可以指定write_only=True,一般默认的可写可读模式就可以了
#    print('所有新建表的默认名字：%s' % wb.sheetnames)
    #直接赋值可以改工作表的名字
    sheet = wb.active
#    sheet.title = '新建表'
    print('所有新建表修改后的名字：%s' % wb.sheetnames)
    #新建一个工作表，表名为‘’，可以指定索引，安排到第二个工作表,index = 0就是第一个位置
#    wb.create_sheet('第二张表',index = 1)
    #删除某一张表
#    wb.remove(sheet)
    #最重要的一步是要保存
    fileName = os.path.join(os.getcwd(),'翻译1.xlsx')
#    print('新建表的路径:' + fileName)
#    wb.save(r'%s' % fileName)
###########################################################################################
#直接给单元格赋值
    sheet['A1'] = 'good'
#B9处写入平均值
    sheet['B9'] = '=AVERAGE(B2:B8)'
#可以一次添加多行数据，从第一行空白行开始（下面都是空白行）写入
    #1.添加一行
    row = [1,2,3,4,5]
    sheet.append(row)
    #2.添加多行
    rows = [
        ['Number', 'data1', 'data2'],
        [2, 40, 30],
        [3, 40, 25],
        [4, 50, 30],
        [5, 30, 10],
        [6, 25, 5],
        [7, 50, 10],
        ]
    for row in rows:
        sheet.append(row)
    #解释下上面的list(zip(*rows))首先*rows将列表打散，相当于填入了若干个参数，zip从某个列表中提取第1个值组合成一个tuple，再从每个列表中提取第2个值组合成一个tuple，一直到最短列表的最后一个值提取完毕后结束，更长列表的之后的值被舍弃，换句话，最后的元组个数是由原来每个参数（可迭代对象）的最短长度决定的。比如现在随便删掉一个值，最短列表长度为2，data2那一列（竖着看）的值全部被舍弃。
    print(list(zip(*rows)))
#设置单元格风格，先导入需要的类from openpyxl.styles import Font, colors, Alignment
    #1、字体
    bold_itatic_24_font = Font(name = '等线',size = 24, italic = True,color = colors.RED,bold = True)
    sheet['A1'].font = bold_itatic_24_font
    sheet['A2'] = 'Test'
#设置A2中的数据垂直居中和水平居中
    sheet['A2'].alignment = Alignment(horizontal = 'center',vertical='center')
#设置行高和列宽
    #1、第2行行高
    sheet.row_dimensions[2].height = 40
    #2、C列列宽
    sheet.column_dimensions['C'].width = 30
    #合并和拆分单元格
    sheet.merge_cells('A13:B14')
    sheet.unmerge_cells('A13:B14')
    wb.save(r'%s' % fileName)

def open_strings_file(filePath):
    for dir in os.listdir(filePath):
        path = os.path.join(filePath,dir)
        if os.path.isdir(path):
            main(path)
        else:
            if os.path.splitext(path)[1] in ['.strings']:
                print(path)

if __name__ == '__main__':
    file_dir_Path = os.getcwd()
#    main(file_dir_Path)
    old_excel = xlrd.open_workbook('iOS翻译汇总.xlsx')
    #读操作：根据下标获取sheet
    #sheet = old_excel.sheet_by_index(0)
    new_excel = copy(old_excel)
    #copy过来的写操作：根据下表获取sheet
    sheet = new_excel.get_sheet(0)
    sheet.write(0,0,'第一行，第一列')
    sheet.write(0,1,'第一行，第二列')
    new_excel.save('new_file.xlsx')

